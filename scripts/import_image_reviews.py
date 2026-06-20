#!/usr/bin/env python3
"""Read reviewed image-description sheet(s) back into the ledger.

For each row, the editor's `alt_text_edited` is recorded against the post's id in
`image_reviews.json`. The generated `alt_text` column is reference only and is
*not* stored — the ledger keeps only human edits plus a lightweight reviewed
marker:

    "facebook:123:456": {"reviewed_at": "…"}                      # approved as-is
    "facebook:123:789": {"reviewed_at": "…", "alt_text_edited": "…"}  # edited

The marker is why a reviewed-but-unedited image doesn't resurface after the next
scrape; the absent text is why the ledger never accumulates copies of generated
descriptions. A blank `alt_text_edited` means "keep the generated text".

Pass a single .xlsx or a directory (all *.xlsx in it are imported — handy for
collecting the split sheets from several reviewers). Rows match by the locked
`id` column; the ledger is stable across re-runs (an unchanged row keeps its
entry and `reviewed_at`), so re-importing produces no diff.
"""

import argparse
from datetime import datetime, timezone
from pathlib import Path

from image_reviews_common import load_ledger, save_ledger
from openpyxl import load_workbook

COL_ID = "id"
COL_EDITED = "alt_text_edited"


def _sheets(path):
    """Resolve a file-or-directory argument to a sorted list of .xlsx paths."""
    path = Path(path)
    if path.is_dir():
        sheets = sorted(path.glob("*.xlsx"))
        if not sheets:
            raise SystemExit(f"No .xlsx files in directory: {path}")
        return sheets
    return [path]


def _read_rows(xlsx):
    """Yield ``(id, alt_text_edited)`` from a review sheet, located by header."""
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    rows = wb.active.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
    except StopIteration:
        raise SystemExit(f"{xlsx} is empty.") from None
    try:
        id_idx = header.index(COL_ID)
        edited_idx = header.index(COL_EDITED)
    except ValueError:
        raise SystemExit(
            f"{xlsx} must have '{COL_ID}' and '{COL_EDITED}' columns; "
            f"found {header}."
        ) from None
    for values in rows:
        key = values[id_idx] if id_idx < len(values) else None
        if key in (None, ""):
            continue
        raw = values[edited_idx] if edited_idx < len(values) else None
        yield str(key).strip(), (str(raw) if raw is not None else "").strip()


def import_reviews(path, ledger_path):
    ledger = load_ledger(ledger_path)
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    changed = unchanged = 0
    for xlsx in _sheets(path):
        for key, edited in _read_rows(xlsx):
            existing = ledger.get(key)
            if existing is not None and existing.get("alt_text_edited", "") == edited:
                unchanged += 1
                continue
            entry = {"reviewed_at": now}
            if edited:
                entry["alt_text_edited"] = edited
            ledger[key] = entry
            changed += 1
    save_ledger(ledger_path, ledger)
    edits = sum(1 for e in ledger.values() if e.get("alt_text_edited"))
    print(
        f"Recorded {changed} review(s) ({unchanged} unchanged) to {ledger_path}; "
        f"ledger now holds {len(ledger)} post(s), {edits} with an edit."
    )
    return changed


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "path", type=Path, help="Reviewed .xlsx, or a directory of them."
    )
    parser.add_argument(
        "--ledger",
        default="scripts/data/image_reviews.json",
        help="Review ledger to update (default: scripts/data/image_reviews.json).",
    )
    args = parser.parse_args()
    import_reviews(args.path, args.ledger)


if __name__ == "__main__":
    main()
