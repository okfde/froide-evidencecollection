#!/usr/bin/env python3
"""Export image descriptions from import.json for human review.

One row per post that has an image (or an existing description): a locked `id`, two
embedded thumbnails (the post's content `image` and its full-page `screenshot`,
each shown whenever present), an
editable `alt_text` cell — pre-filled with the machine-generated description, or
left blank for an image that has none so an editor can write one by hand — and a
free-text `notes` cell. A reviewer edits `alt_text` in place (or clears it to drop
a description) and the sheet goes back in via import_image_reviews.py.

By default only posts not yet in the ledger are exported, so a freshly prepared
import.json surfaces only its new images; pass --all to re-export everything for
a re-check pass. --split N divides the rows across N sheets for parallel review;
each sheet is self-contained (import_image_reviews matches rows by id).

Thumbnails are read from --images-root, the import bundle's root directory (the
folder the posts' "./images/…" and "./screenshots/…" paths resolve against).
"""

import argparse
import os
import sys
from io import BytesIO
from pathlib import Path

from image_reviews_common import (
    ensure_parent,
    generated_alt,
    image_path,
    iter_review_posts,
    load_json,
    load_ledger,
    screenshot_path,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from PIL import Image as PILImage

# Generated text is shown read-only as a reference; the editor writes into
# `alt_text_edited`. Keep these names in sync with import_image_reviews.py.
HEADERS = ["id", "image", "full_screenshot", "alt_text", "alt_text_edited", "notes"]
THUMB_MAX = (320, 320)
# The full-page screenshot carries more detail (surrounding text, layout), and is
# usually tall/portrait — so cap it by width (what makes text legible) and leave a
# generous height ceiling, letting tall shots render at a readable width rather
# than being shrunk to fit a square box.
SCREENSHOT_MAX = (480, 800)
PX_PER_CHAR = 7.0  # Excel column-width unit ≈ 7px at the default font.
PX_TO_POINT = 0.75  # Row height is in points.


def _pending(data, ledger, include_all):
    """Collect ``(key, generated_text, prior_edit, item)`` for rows to review.

    `prior_edit` pre-fills the editable column on a re-review pass (--all); it is
    empty for first-time rows. A post can be grouped under more than one scrape
    target, so the same key appears multiple times in import.json; emit each
    only once."""
    rows = []
    seen = set()
    for key, _platform, item in iter_review_posts(data):
        if key in seen:
            continue
        if include_all or key not in ledger:
            prior_edit = (ledger.get(key) or {}).get("alt_text_edited", "")
            rows.append((key, generated_alt(item), prior_edit, item))
            seen.add(key)
    return rows


def _shards(out, rows, split):
    """Yield ``(path, rows_chunk)``. split=1 keeps `out` as given; otherwise
    contiguous balanced chunks land in ``<stem>_NN<suffix>``."""
    if split == 1:
        yield out, rows
        return
    stem, suffix = os.path.splitext(out)
    suffix = suffix or ".xlsx"
    base, extra = divmod(len(rows), split)
    start = 0
    for i in range(split):
        size = base + (1 if i < extra else 0)
        if size == 0:
            continue
        yield f"{stem}_{i + 1:02d}{suffix}", rows[start : start + size]
        start += size


def _embed_thumbnail(ws, anchor, path, buffers, max_size=THUMB_MAX):
    """Embed a scaled thumbnail at `anchor` (e.g. "B2"); return display height in
    px, or None if there's no image to embed."""
    if path is None:
        return None
    try:
        with PILImage.open(path) as im:
            im = im.convert("RGB")
            im.thumbnail(max_size)
            buf = BytesIO()
            im.save(buf, format="PNG")
            w, h = im.size
    except Exception as exc:  # noqa: BLE001 — a bad file shouldn't abort the export
        print(f"Could not read {path}: {exc}", file=sys.stderr)
        return None
    buf.seek(0)
    buffers.append(buf)  # keep alive until the workbook is saved
    xl = XLImage(buf)
    xl.width, xl.height = w, h
    ws.add_image(xl, anchor)
    return h


def _write_workbook(path, rows, images_root):
    """Build and save one review workbook; return (rows, missing-image count)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "image_reviews"

    bold = Font(bold=True)
    for col, name in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=name).font = bold
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = THUMB_MAX[0] / PX_PER_CHAR
    ws.column_dimensions["C"].width = SCREENSHOT_MAX[0] / PX_PER_CHAR
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 30

    wrap = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    buffers = []
    missing = 0

    for r, (key, generated, prior_edit, item) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=key).alignment = top
        ws.cell(row=r, column=4, value=generated).alignment = wrap
        ws.cell(row=r, column=5, value=prior_edit or None).alignment = wrap
        ws.cell(row=r, column=6).alignment = wrap
        img_h = _embed_thumbnail(ws, f"B{r}", image_path(item, images_root), buffers)
        shot_h = _embed_thumbnail(
            ws, f"C{r}", screenshot_path(item, images_root), buffers, SCREENSHOT_MAX
        )
        if img_h is None:
            ws.cell(row=r, column=2, value="[no image]").alignment = wrap
        if shot_h is None:
            ws.cell(row=r, column=3, value="[no screenshot]").alignment = wrap
        heights = [h for h in (img_h, shot_h) if h is not None]
        if heights:
            ws.row_dimensions[r].height = max(max(heights) * PX_TO_POINT, 60)
        else:
            missing += 1  # no visual at all (neither image nor screenshot)
            ws.row_dimensions[r].height = 60

    wb.save(ensure_parent(path))
    return len(rows), missing


def export_reviews(import_json, images_root, out, ledger_path, split, include_all):
    if not os.path.isdir(images_root):
        raise SystemExit(f"--images-root is not a directory: {images_root}")
    if split < 1:
        raise SystemExit("--split must be at least 1.")

    data = load_json(import_json)
    ledger = load_ledger(ledger_path)
    rows = _pending(data, ledger, include_all)
    if not rows:
        print("No image descriptions to review.")
        return []

    written = []
    for path, chunk in _shards(out, rows, split):
        total, missing = _write_workbook(path, chunk, images_root)
        written.append(path)
        print(f"Wrote {total} row(s) to {path} ({missing} without a thumbnail).")

    # Images carrying no description yet — the rows an editor must caption by
    # hand (blank generated text and no prior edit). Report count and ids.
    no_alt = [
        key
        for key, generated, prior_edit, _item in rows
        if not (generated or prior_edit)
    ]
    if no_alt:
        print(f"{len(no_alt)} of {len(rows)} image(s) have no alt_text yet:")
        for key in no_alt:
            print(f"  {key}")
    else:
        print(f"All {len(rows)} exported image(s) already have alt_text.")
    return written


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("import_json", type=Path, help="Path to import.json.")
    parser.add_argument(
        "--images-root",
        required=True,
        help="Import bundle root the image/screenshot paths resolve against.",
    )
    parser.add_argument(
        "--out", default="scripts/data/reviews.xlsx", help="Output XLSX path."
    )
    parser.add_argument(
        "--ledger",
        default="scripts/data/image_reviews.json",
        help="Review ledger; reviewed posts are skipped "
        "(default: scripts/data/image_reviews.json).",
    )
    parser.add_argument(
        "--split",
        type=int,
        default=1,
        help="Split rows across N sheets (<out>_01.xlsx …) for parallel review.",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Include already-reviewed posts (default: only un-reviewed).",
    )
    args = parser.parse_args()
    export_reviews(
        args.import_json,
        args.images_root,
        args.out,
        args.ledger,
        args.split,
        args.all,
    )


if __name__ == "__main__":
    main()
