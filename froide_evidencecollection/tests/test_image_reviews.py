"""Round-trip tests for the file-based image-review scripts in scripts/.

These tools never touch the database; they operate on import.json, a ledger, and
image files. The scripts directory is added to the path so the modules import
like they do when run as `python scripts/<name>.py`.
"""

import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image as PILImage

SCRIPTS = Path(__file__).resolve().parents[2] / "scripts"
sys.path.insert(0, str(SCRIPTS))

import apply_image_reviews  # noqa: E402
import export_image_reviews  # noqa: E402
import import_image_reviews  # noqa: E402


def _post(post_id, user_id, image_file=None, screenshot=None, alt="generated"):
    item = {
        "platform_post_id": post_id,
        "url": f"https://x/{post_id}",
        "account": {"platform_user_id": user_id, "username": "acct"},
    }
    if alt is not None:  # alt=None models an image with no description at all
        item["image_alt_text"] = {"text_bezug_zum_bild": "JA", "alt_text": alt}
    if image_file:
        item["image_file"] = image_file
    if screenshot:
        item["screenshot_file"] = screenshot
    return item


def _bundle(tmp_path, posts_by_platform):
    """Write an import.json shaped like prepare_import's output; return its path."""
    data = {"hash1": {"social_media": posts_by_platform}}
    path = tmp_path / "import.json"
    path.write_text(json.dumps(data))
    return path


def _make_image(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    PILImage.new("RGB", (40, 25), "blue").save(path)


def test_export_import_apply_roundtrip(tmp_path):
    _make_image(tmp_path / "images" / "a.jpg")
    _make_image(tmp_path / "screenshots" / "b.png")
    import_json = _bundle(
        tmp_path,
        {
            "facebook": [
                _post("1", "u1", image_file="./images/a.jpg"),
                _post(
                    "2", "u1", screenshot="./screenshots/b.png"
                ),  # screenshot fallback
            ],
            "twitter": [
                _post("3", "u9", alt=""),  # no generated text -> not exported
            ],
        },
    )
    ledger = tmp_path / "image_reviews.json"
    out = tmp_path / "reviews.xlsx"

    written = export_image_reviews.export_reviews(
        import_json, str(tmp_path), str(out), str(ledger), split=1, include_all=False
    )
    assert written == [str(out)]

    ws = load_workbook(out).active
    assert [c.value for c in ws[1]] == [
        "id",
        "image",
        "alt_text",
        "alt_text_edited",
        "notes",
    ]
    keys = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert keys == {"facebook:u1:1", "facebook:u1:2"}  # post 3 excluded

    # Editor corrects post 1 (column 4); leaves post 2 untouched (approve as-is).
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "facebook:u1:1":
            ws.cell(row=r, column=4, value="corrected")
    ws.parent.save(out)

    import_image_reviews.import_reviews(str(out), str(ledger))
    led = json.loads(ledger.read_text())
    assert led["facebook:u1:1"]["alt_text_edited"] == "corrected"
    assert "alt_text_edited" not in led["facebook:u1:2"]  # approved, no edit stored
    assert "alt_text" not in led["facebook:u1:1"]  # generated text not copied in
    assert "reviewed_at" in led["facebook:u1:2"]

    # Apply overrides only the edited post; the approved one keeps generated text.
    curated = tmp_path / "import.curated.json"
    apply_image_reviews.apply_reviews(import_json, str(ledger), str(curated))
    data = json.loads(curated.read_text())
    posts = {
        p["platform_post_id"]: p for p in data["hash1"]["social_media"]["facebook"]
    }
    assert posts["1"]["image_alt_text"]["alt_text"] == "corrected"
    assert posts["1"]["image_alt_text"]["text_bezug_zum_bild"] == "JA"
    assert posts["2"]["image_alt_text"]["alt_text"] == "generated"  # unchanged

    # Reviewed posts drop out of the next export; nothing left -> no file.
    written2 = export_image_reviews.export_reviews(
        import_json,
        str(tmp_path),
        str(tmp_path / "reviews2.xlsx"),
        str(ledger),
        split=1,
        include_all=False,
    )
    assert written2 == []
    assert not (tmp_path / "reviews2.xlsx").exists()


def test_image_without_description_is_exported_for_handfilling(tmp_path):
    _make_image(tmp_path / "images" / "a.jpg")
    import_json = _bundle(
        tmp_path,
        {
            "facebook": [
                _post(
                    "1", "u1", image_file="./images/a.jpg", alt=None
                ),  # image, no desc
                _post("2", "u1", alt=""),  # no image, no desc -> excluded
            ]
        },
    )
    ledger = tmp_path / "image_reviews.json"
    out = tmp_path / "reviews.xlsx"
    export_image_reviews.export_reviews(
        import_json, str(tmp_path), str(out), str(ledger), 1, False
    )
    ws = load_workbook(out).active
    rows = {
        ws.cell(row=r, column=1).value: (
            ws.cell(row=r, column=3).value,  # generated (reference)
            ws.cell(row=r, column=4).value,  # editable
        )
        for r in range(2, ws.max_row + 1)
    }
    assert set(rows) == {"facebook:u1:1"}  # only the image post; post 2 excluded
    assert rows["facebook:u1:1"] == (None, None)  # no generated text, blank to fill

    # Editor writes a description by hand (column 4) for the image that had none.
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "facebook:u1:1":
            ws.cell(row=r, column=4, value="hand written")
    ws.parent.save(out)
    import_image_reviews.import_reviews(str(out), str(ledger))

    curated = tmp_path / "import.curated.json"
    apply_image_reviews.apply_reviews(import_json, str(ledger), str(curated))
    data = json.loads(curated.read_text())
    post = data["hash1"]["social_media"]["facebook"][0]
    assert post["image_alt_text"]["alt_text"] == "hand written"


def test_export_reports_images_without_alt_text(tmp_path, capsys):
    _make_image(tmp_path / "images" / "a.jpg")
    import_json = _bundle(
        tmp_path,
        {
            "facebook": [
                _post("1", "u1", image_file="./images/a.jpg"),  # has description
                _post("2", "u1", image_file="./images/a.jpg", alt=None),  # none
            ]
        },
    )
    export_image_reviews.export_reviews(
        import_json,
        str(tmp_path),
        str(tmp_path / "r.xlsx"),
        str(tmp_path / "led.json"),
        1,
        False,
    )
    out = capsys.readouterr().out
    assert "1 of 2 image(s) have no alt_text yet:" in out
    listing = out.split("no alt_text yet:")[1]
    assert "facebook:u1:2" in listing  # the captionless image is named
    assert "facebook:u1:1" not in listing  # the described one is not


def test_only_new_posts_surface_after_rescrape(tmp_path):
    _make_image(tmp_path / "images" / "a.jpg")
    import_json = _bundle(
        tmp_path, {"facebook": [_post("1", "u1", image_file="./images/a.jpg")]}
    )
    ledger = tmp_path / "image_reviews.json"
    out = tmp_path / "reviews.xlsx"
    export_image_reviews.export_reviews(
        import_json, str(tmp_path), str(out), str(ledger), 1, False
    )
    import_image_reviews.import_reviews(str(out), str(ledger))

    # A re-scrape regenerates import.json with the old post plus a new one.
    import_json = _bundle(
        tmp_path,
        {
            "facebook": [
                _post("1", "u1", image_file="./images/a.jpg"),
                _post("2", "u1", image_file="./images/a.jpg"),
            ]
        },
    )
    out2 = tmp_path / "reviews_new.xlsx"
    export_image_reviews.export_reviews(
        import_json, str(tmp_path), str(out2), str(ledger), 1, False
    )
    ws = load_workbook(out2).active
    keys = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert keys == {"facebook:u1:2"}  # only the new post


def test_split_produces_disjoint_sheets(tmp_path):
    _make_image(tmp_path / "images" / "a.jpg")
    posts = [_post(str(i), "u1", image_file="./images/a.jpg") for i in range(5)]
    import_json = _bundle(tmp_path, {"facebook": posts})
    ledger = tmp_path / "image_reviews.json"

    written = export_image_reviews.export_reviews(
        import_json, str(tmp_path), str(tmp_path / "r.xlsx"), str(ledger), 2, False
    )
    assert written == [str(tmp_path / "r_01.xlsx"), str(tmp_path / "r_02.xlsx")]

    def keys(path):
        ws = load_workbook(path).active
        return {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}

    k1, k2 = keys(written[0]), keys(written[1])
    assert len(k1) == 3 and len(k2) == 2 and k1.isdisjoint(k2)


def test_import_directory_collects_all_sheets(tmp_path):
    _make_image(tmp_path / "images" / "a.jpg")
    posts = [_post(str(i), "u1", image_file="./images/a.jpg") for i in range(4)]
    import_json = _bundle(tmp_path, {"facebook": posts})
    ledger = tmp_path / "image_reviews.json"
    sheets_dir = tmp_path / "sheets"
    sheets_dir.mkdir()

    # Two reviewers each get a sheet; each edits one row.
    export_image_reviews.export_reviews(
        import_json, str(tmp_path), str(sheets_dir / "r.xlsx"), str(ledger), 2, False
    )
    for sheet, text in [
        (sheets_dir / "r_01.xlsx", "edit-a"),
        (sheets_dir / "r_02.xlsx", "edit-b"),
    ]:
        wb = load_workbook(sheet)
        wb.active.cell(row=2, column=4, value=text)
        wb.save(sheet)

    # Importing the directory picks up both sheets at once.
    import_image_reviews.import_reviews(str(sheets_dir), str(ledger))
    led = json.loads(ledger.read_text())
    assert len(led) == 4  # all four reviewed
    edits = sorted(e["alt_text_edited"] for e in led.values() if "alt_text_edited" in e)
    assert edits == ["edit-a", "edit-b"]


@pytest.mark.parametrize("rescrape_changes_nothing", [True])
def test_unchanged_reimport_keeps_reviewed_at(tmp_path, rescrape_changes_nothing):
    _make_image(tmp_path / "images" / "a.jpg")
    import_json = _bundle(
        tmp_path, {"facebook": [_post("1", "u1", image_file="./images/a.jpg")]}
    )
    ledger = tmp_path / "image_reviews.json"
    out = tmp_path / "reviews.xlsx"
    export_image_reviews.export_reviews(
        import_json, str(tmp_path), str(out), str(ledger), 1, False
    )
    import_image_reviews.import_reviews(str(out), str(ledger))
    first = json.loads(ledger.read_text())["facebook:u1:1"]["reviewed_at"]

    # Re-importing the same (unchanged) sheet must not churn the timestamp.
    import_image_reviews.import_reviews(str(out), str(ledger))
    second = json.loads(ledger.read_text())["facebook:u1:1"]["reviewed_at"]
    assert first == second
